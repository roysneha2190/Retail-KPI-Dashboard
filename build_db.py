import sqlite3
import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

db_path = "RetailIntelligence.db"
excel_path = "Retail_Dashboard_Data.xlsx"

try:
    conn = sqlite3.connect(db_path)

    # 1. Pull Overall Summary KPIs
    query_kpis = """
    SELECT 
        COUNT(*) AS Total_Orders,
        ROUND(SUM(Total_Sales), 2) AS Total_Revenue,
        ROUND(SUM(Profit), 2) AS Total_Profit,
        ROUND((SUM(Profit) / SUM(Total_Sales)) * 100, 2) AS Profit_Margin_Percent
    FROM sales;
    """
    df_kpis = pd.read_sql_query(query_kpis, conn)

    # 2. Pull Category Breakdown
    query_categories = """
    SELECT 
        Product_Category, 
        ROUND(SUM(Total_Sales), 2) AS Revenue, 
        ROUND(SUM(Profit), 2) AS Profit
    FROM sales 
    GROUP BY Product_Category 
    ORDER BY Revenue DESC;
    """
    df_categories = pd.read_sql_query(query_categories, conn)

    # 3. New: Pull Top 5 Countries by Sales Volume
    query_countries = """
    SELECT 
        Country, 
        ROUND(SUM(Total_Sales), 2) AS Total_Sales
    FROM sales 
    GROUP BY Country 
    ORDER BY Total_Sales DESC 
    LIMIT 5;
    """
    df_countries = pd.read_sql_query(query_countries, conn)
    conn.close()

    # 4. Save raw data to separate sheets
    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        df_kpis.to_excel(writer, sheet_name="Overall_KPIs", index=False)
        df_categories.to_excel(writer, sheet_name="Categories", index=False)
        df_countries.to_excel(writer, sheet_name="Countries", index=False)

    # 5. Open workbook to build the layout
    wb = load_workbook(excel_path)
    ws_dash = wb.create_sheet(title="Executive_Dashboard", index=0)
    ws_dash.views.sheetView[0].showGridLines = True

    # Styling Assets
    navy_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
    ice_blue_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=11, bold=True)
    title_font = Font(name="Calibri", size=16, bold=True, color="1F497D")
    
    thin_border = Border(
        left=Side(style='thin', color='B0C4DE'), right=Side(style='thin', color='B0C4DE'),
        top=Side(style='thin', color='B0C4DE'), bottom=Side(style='thin', color='B0C4DE')
    )

    # Add Dashboard Title
    ws_dash["A1"] = "RETAIL INTELLIGENCE EXECUTIVE DASHBOARD"
    ws_dash["A1"].font = title_font

    # Populate KPI Cards
    headers = ["Total Orders", "Total Revenue", "Total Profit", "Profit Margin"]
    values = [df_kpis.iloc[0]['Total_Orders'], df_kpis.iloc[0]['Total_Revenue'], 
              df_kpis.iloc[0]['Total_Profit'], f"{df_kpis.iloc[0]['Profit_Margin_Percent']}%"]
    
    for i, (head, val) in enumerate(zip(headers, values), start=1):
        col_letter = chr(64 + i)
        cell_hdr = ws_dash[f"{col_letter}3"]
        cell_hdr.value = head
        cell_hdr.fill = navy_fill
        cell_hdr.font = white_font
        cell_hdr.alignment = Alignment(horizontal="center")
        cell_hdr.border = thin_border
        
        cell_val = ws_dash[f"{col_letter}4"]
        cell_val.value = val
        cell_val.fill = ice_blue_fill
        cell_val.font = bold_font
        cell_val.alignment = Alignment(horizontal="center")
        cell_val.border = thin_border
        
        if i in [2, 3]:
            cell_val.number_format = '$#,##0.00'

    # 6. Re-create the Category Bar Chart
    ws_cats = wb["Categories"]
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Revenue vs Profit by Category"
    bar_chart.y_axis.title = "Financials ($)"
    bar_chart.x_axis.title = "Product Category"
    
    max_cat_row = ws_cats.max_row
    bar_data = Reference(ws_cats, min_col=2, min_row=1, max_col=3, max_row=max_cat_row)
    bar_cats = Reference(ws_cats, min_col=1, min_row=2, max_row=max_cat_row)
    bar_chart.add_data(bar_data, titles_from_data=True)
    bar_chart.set_categories(bar_cats)
    
    ws_dash.add_chart(bar_chart, "A7")
    bar_chart.width = 14
    bar_chart.height = 10

    # 7. Create the New Country Share Doughnut Chart
    ws_countries = wb["Countries"]
    pie_chart = PieChart()
    pie_chart.holeSize = 50  # Turns a standard pie chart into a modern doughnut style
    pie_chart.title = "Top 5 Global Markets Share"
    pie_chart.style = 10
    
    max_count_row = ws_countries.max_row
    pie_data = Reference(ws_countries, min_col=2, min_row=1, max_row=max_count_row)
    pie_cats = Reference(ws_countries, min_col=1, min_row=2, max_row=max_count_row)
    pie_chart.add_data(pie_data, titles_from_data=True)
    pie_chart.set_categories(pie_cats)
    
    # Place this chart directly to the right of the bar chart (Column I)
    ws_dash.add_chart(pie_chart, "I7")
    pie_chart.width = 12
    pie_chart.height = 10

    # Fix Column Sizes
    for col in ws_dash.columns:
        ws_dash.column_dimensions[col[0].column_letter].width = 18

    wb.save(excel_path)
    print("\n🎉 SUCCESS! Country Share Doughnut Chart added to your dashboard layout!")

except Exception as e:
    print(f"\n❌ Error: {e}")